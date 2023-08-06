"""Read EMA data stored in an Excel 2007- (xlsx) workbook file.

Each Excel work-book file, and each work-sheet in the file,
may include EMA records for
one or more subject(s), each reporting on one or more EMA variables,
representing nominal SCENARIO categories and/or ordinal ATTRIBUTE ratings.

Each row in the worksheet must include enough data to identify
all requested variables.

*** Usage Example:
    See module ema_data and main template script run_ema

*** Version History:
* Version 0.5.1:
2021-11-26, handle missing file field nicely

* Version 0.5:
2021-10-20, first functional version
2021-11-17, save method
"""
# ********** check file exceptions **************
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
# because Openpyxl might try to open bad xlsx file as a zip archive
from openpyxl.utils import column_index_from_string
import logging

from EmaCalc import ema_file

logger = logging.getLogger(__name__)


class FileFormatError(ema_file.FileReadError):
    """Format error causing non-usable data"""


class ParameterError(ema_file.FileReadError):
    """Error in calling parameters causing non-usable data in some file."""


class ArgumentError(RuntimeError):
    """Error in calling arguments. No file can be read."""


class EmaFile(ema_file.EmaFile):
    """Interface to Excel xlsx workbook file storing count-profile data.
    Each sheet in the workbook may include data for
    one or more subjects, count factors.
    Data elements MUST be stored at the same location in all sheets.
    """
    def __init__(self, file_path,
                 ema_vars,
                 top_row=2,
                 sheets=None,
                 subject=None,
                 ):
        """File interface to data stored in an excel file,
        decoded as a sequence of PairedCompItem instances.
        :param file_path: Path to file for reading
        :param top_row: integer address of first row containing PairedCompItem data
        :param sheets: (optional) list of sheet names to be searched for data.
        :param subject: 'sheet' or column for subject identification label
            **** allow subject = 'file' name ? *******
        :param ema_vars: dict with elements (ema_key, location), where
            ema_key is a string identifying one count key,
            location is a string with a column (top) address like 'D'.
        """
        super().__init__(file_path)
        self.subject = _check_column_or_sheet(subject)
        self.ema_vars = ema_vars
        self.sheets = sheets
        self.top_row = top_row
        # self.test_factors = _check_test_cond(test_factors)

    def __iter__(self):
        """Generator yielding data from an excel file.
        :return: generator yielding EmaRecord instances
        """
        try:
            wb = load_workbook(str(self.file_path), read_only=True)
        except Exception as e:  # openpyxl might raise InvalidFileException, BadZipFile, other?
            raise FileFormatError(f'Cannot load Excel workbook from file {self.file_path.stem}. '
                                  + f'Openpyxl error: {e}')
        if self.sheets is None:
            sheets = wb.sheetnames
        else:
            sheets = set(self.sheets) & set(wb.sheetnames)
        if len(sheets) == 0:
            raise FileFormatError(f'No accepted sheets found in {self.file_path.stem}')
        for sheet_name in sheets:
            ws = wb[sheet_name]
            rows = ws.rows
            for _ in range(self.top_row - 1):
                try:
                    row = rows.__next__()
                    logger.debug(f'skipping row {row[0].row}')
                except StopIteration:
                    raise FileFormatError(f'No data rows in {self.file_path.stem}.'
                                          + f'sheet {repr(sheet_name)}')
            for row in rows:
                r = ema_file.EmaRecord(subject=self._get_subject(ws, row),
                                       ema=self._get_ema_dict(ws, row)
                                       )
                if r.subject is not None:
                    # or any(r_i is None  # **** let caller check this!
                    #      for r_i in r.ema.values())):
                    # we have a valid record
                    yield r  # ******
                else:
                    logger.debug(f'not using row {row[0]}...')

    def _get_subject(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: subject code, integer or string
        """
        return _get_value(ws, self.subject, row)

    def _get_ema_dict(self, ws, row):
        """
        :param ws: a worksheet
        :param row: tuple of openpyxl Cell instances
        :return: dict with tuples (ema_key, category_label), where
            ema_key is string defining an EMA 'dimension'
            category_label = numeric or string label of response category in this factor
        """
        return dict([(cf, _get_value(ws, c_col, row))
                     for (cf, c_col) in self.ema_vars.items()])

    # ----------------------------------------------- save data to workbook
    def save(self, subject, items, allow_over_write=False):
        """Save EMA records to file with ONE sheet for ONE subject,
        identified in the sheet title.
        :param subject: string with subject id
        :param items: iterable of dicts with EMA data
        :param allow_over_write: boolean switch, =True allows old file to be over-written
        :return: None

        Result: a new file is created.

        2021-11-17, new method for this file format
        """
        # *** update self properties to allow reading from the file?
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=subject)
        header = self._make_header()
        n_col = len(header)
        ws.append(header)
        self.top_row = 2
        for ema in items:
            ws.append(self._make_row(ema, n_col))
        if not allow_over_write:
            self.file_path = ema_file.safe_file_path(self.file_path)
        self.sheets = wb.sheetnames
        wb.save(self.file_path)

    def _make_header(self):  # **** module function or static ******
        """Make header list from column addresses of self.
        :param n_col:
        :return: list of header labels, length = n_col
        """
        n_col = max(column_index_from_string(v)
                    for v in self.ema_vars.values())
        h = [None for _ in range(n_col)]
        for (k, c) in self.ema_vars.items():
            h[column_index_from_string(c) - 1] = k
        return h

    def _make_row(self, ema, n_col):
        """Make list with contents from given ema record
        :param ema: a dict with EMA data
        :return: r = list filled with elements from ema,
            placed at columns defined by self.ema_vars
        """
        r = [None for _ in range(n_col)]
        for (k, v) in ema.items():
            if k in self.ema_vars.keys():
                r[column_index_from_string(self.ema_vars[k]) - 1] = v
        return r


# --------------------------------------------------- help sub-functions

def _check_column(col):
    """Check that a parameter address is acceptable
    """
    if col is None or type(col) is str and col.isupper():
        return col
    # otherwise:
    raise ArgumentError(f'Column address {col} must be string of uppercase letters')


def _check_column_or_sheet(col):
    """Check that a parameter address is either 'sheet' or a column string address
    """
    if type(col) is str and col == 'sheet':
        return col
    else:
        return _check_column(col)


def _get_value(ws, col, row):
    """Get contents in ONE cell or in sheet title
    :param ws: a worksheet
    :param col: one column address or 'sheet'
    :param row: integer address of row
    :return: cell contents
    """
    if col == 'sheet':
        return ws.title
    # c = ws[col + str(row)]
    try:
        cell = row[column_index_from_string(col) - 1]
        return cell.value
    except IndexError:
        return None  # OR raise ArgumentError ?

