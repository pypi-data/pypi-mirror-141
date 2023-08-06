from inspect import indentsize
from openpyxl import load_workbook
import json
import os


def parse_cabinet(excel_file):
    collection = excel_file
    file = os.path.join(os.path.dirname(__file__),
                        'data', f"{excel_file}.xlsx")
    ws = load_workbook(file).active
    minrow = int(ws.min_row)
    maxrow = int(ws.max_row)
    lst = []
    retdict = {}
    for row in ws.iter_rows(minrow, maxrow):
        for cell in row:
            if cell.value is not None:
                if "KLEURTHEMA" in cell.value:
                    lst.append(int(cell.row))

    for index in range(len(lst)):
        max = 0
        min = lst[index]
        max = lst[index + 1] if index != len(lst)-1 else maxrow
        for index, row in enumerate(ws.iter_rows(min, max)):
            # print("-------------------------------------------------------")
            for cell in row:
                if cell.value is not None:
                    if "KLEURTHEMA" in cell.value:
                        color_theme = val_to_color_theme(cell.value)
                    elif "BEHANG" in cell.value:
                        product_type = "wallpaper"
                    elif "VERF" in cell.value:
                        product_type = "paint"
                    elif "STOF" in cell.value:
                        product_type = "fabric"
                    else:
                        name = ws.cell(row=min+index, column=7).value
                        code = ws.cell(row=min+index, column=6).value
                        #print(cell, cell.value)
            if name is not None and code is not None:
                retdict[code] = {
                    'name': val_to_name(name),
                    'code': code,
                    'color_theme': color_theme,
                    'product_type': product_type,
                    'collection': collection
                }
    return retdict


def parse_excel(excel_file):
    collection = excel_file
    # print(collection)
    file = os.path.join(os.path.dirname(__file__),
                        'data', f"{excel_file}.xlsx")
    # print(file)
    ws = load_workbook(file).active
    minrow = int(ws.min_row)
    maxrow = int(ws.max_row)
    lst = []
    for row in ws.iter_rows(minrow, maxrow):
        for cell in row:
            if cell.value is not None:
                # print(cell.value)
                if "KLEURTHEMA" in cell.value:
                    lst.append(int(cell.row))

    color_theme = None
    product_type = None
    retdict = {}

    for index in range(len(lst)):
        max = 0
        min = lst[index]
        max = lst[index + 1] if index != len(lst)-1 else maxrow
        for index, row in enumerate(ws.iter_rows(min, max)):
            name = None
            code = None
            for cell in row:
                if cell.value is not None:
                    if "KLEURTHEMA" in cell.value:
                        color_theme = val_to_color_theme(cell.value)
                    elif "BEHANG" in cell.value:
                        product_type = "wallpaper"
                    elif "VERF" in cell.value:
                        product_type = "paint"
                    elif "STOF" in cell.value:
                        product_type = "fabric"
                    else:
                        name = ws.cell(row=min+index, column=3).value
                        code = ws.cell(row=min+index, column=7).value
            # print(name)
            # print(code)
            if name is not None and code is not None:
                #print(code, name, color_theme, product_type)
                retdict[code] = {
                    'name': val_to_name(name),
                    'code': code,
                    'color_theme': color_theme,
                    'product_type': product_type,
                    'collection': collection
                }
    return retdict


def val_to_color_theme(value):
    return value[13:].strip().lower().replace(' ', '-')


def val_to_name(value):
    return value.strip().lower().replace(' ', '-')


if __name__ == '__main__':

    bigdict = {}
    bigdict.update(parse_excel('kimono'))
    bigdict.update(parse_excel('orbital'))
    bigdict.update(parse_excel('agathe'))
    bigdict.update(parse_cabinet('cabinet'))
    for product in bigdict.values():
        print(json.dumps(product, indent=2))
