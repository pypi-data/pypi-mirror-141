"""Configuration for the Excel interface."""

from openpyxl.styles import Font, NamedStyle, PatternFill


# File suffixes.
# Cf. https://www.quora.com/What-is-the-extension-of-Excel-files
WORKBOOK_SUFFIX = '.xlsx'

# Values.
# Optimoptions.
AUTO_STARTDATE_VALUES = ('auto', 'automatic')

# Help tab.
HELP_SECTIONHEADER_WIDTH = 8  # cells
HELP_TAB_COLOR = 'de781f'

# NamedStyles.
# header
header = NamedStyle(name='header')
header.font = Font(color='ffffff', bold=True)
header.fill = PatternFill(fill_type='solid', start_color='a0a0a0')
# help_sectionheader
help_sectionheader = NamedStyle(name='help_sectionheader')
help_sectionheader.font = Font(bold=True)
help_sectionheader.fill = PatternFill(fill_type='solid', start_color='dfebe4')

STYLES = (header, help_sectionheader)
