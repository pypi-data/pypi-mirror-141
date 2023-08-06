"""Functionality for the Excel interface.

Class definitions:
BaseTemplateHandler
ProjectTemplateHandler
TestDataTemplateHandler
BaseTemplateWriter
ProjectTemplateWriter
TestDataTemplateWriter
BaseTemplateReader
ProjectTemplateReader
TestDataTemplateReader
"""

import json
import pathlib

import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from eagers.config.dataset import TIMESTAMP_RESOLUTION_TABLENAMES
from eagers.config.excel_interface import (
    AUTO_STARTDATE_VALUES, HELP_SECTIONHEADER_WIDTH, HELP_TAB_COLOR, STYLES,
    WORKBOOK_SUFFIX)
from eagers.config.path_spec import (
    EPW_SUFFIX, HDF5_SUFFIX, EXCEL_INTERFACE_PROJECT_HELP, USER_DIR_HDF5_DATASETS, USER_DIR_LOCATION,
    USER_DIR_PROJECTS, USER_DIR_TESTDATA,BUILDPLUS_DIR_PATH,
)
from eagers.config.text import RUNTIME_HDF5_FILE_PREFIX
from eagers.extras import bplus_load_building
from eagers.read.example_file import load_example_file
from eagers.read.weather_file import read_weather_file
from eagers.basic.datetime_ext import datetime_from_data
from eagers.basic.file_handling import ensure_suffix
from eagers.basic.hdf5 import DatetimeFloatConverter as DFC
from eagers.basic.hdf5 import  new_hdf5_file, write_records_to_new_hdf5_table, write_user_attributes


class BaseTemplateHandler:
    """Defines functionality common between BaseTemplateWriter and
    BaseTemplateReader.
    """

    # Subclass properties.
    _default_filename: str
    _save_dir: pathlib.Path

    _sheet_properties = {}

    def __init__(self):
        self._filename = self._default_filename

    @property
    def filename(self):
        return ensure_suffix(self._filename, WORKBOOK_SUFFIX)

    @filename.setter
    def filename(self, value):
        if value:
            self._filename = value
        else:
            self._filename = self._default_filename

    @property
    def filepath(self) -> pathlib.Path:
        return self._save_dir / self.filename


class ProjectTemplateHandler(BaseTemplateHandler):
    _default_filename = 'project_template'
    _save_dir = USER_DIR_PROJECTS


class TestDataTemplateHandler(BaseTemplateHandler):
    _default_filename = 'testdata_template'
    _save_dir = USER_DIR_TESTDATA


class BaseTemplateWriter(BaseTemplateHandler):
    """Base template writer class."""

    _sheet_order = []

    def write(self, filename=None):
        """Write template to file.

        Keyword arguments:
        filename - (str) (Default: None) Custom file name.
        """
        # Create new workbook.
        wb = Workbook()

        # Add styles.
        for style in STYLES:
            wb.add_named_style(style)

        # Add sheets.
        for i, sheet_name in enumerate(self._sheet_order):
            self._add_sheet(wb, sheet_name, i)

        # Save template file.
        # Use property magic to update the filepath by setting the
        # filename.
        self.filename = filename
        wb.save(self.filepath)

    @classmethod
    def write_template(cls, filename=None):
        """Convenience method that requires no instantiation. Call
        signature is identical to write().
        """
        writer = cls()
        return writer.write(filename)

    @classmethod
    def _add_sheet(cls, wb, name, idx):
        """Generalized sheet creation operations.

        Positional arguments:
        wb - (Workbook) Workbook to add sheet to.
        name - (str) Name of sheet to add.
        idx - (int) Index of the new sheet.
        """
        # Generic operations.
        if idx:
            ws = wb.create_sheet(name)
        else:
            ws = wb.active
            ws.title = name
        ws.sheet_properties.tabColor = cls._sheet_properties[name].get('tab_color')
        # Call method for specific operations.
        return getattr(cls, cls._sheet_properties[name]['add_method'])(ws)


class ProjectTemplateWriter(ProjectTemplateHandler, BaseTemplateWriter):
    """Project template writer class."""

    # Sheet order makes up for the lack of ordering in dictionaries
    # (changed in Python 3.7: dict class now remembers insertion order).
    _sheet_order = ['Components', 'Buildings', 'FluidLoops', 'Network','Options']
    _sheet_properties = {
        'Components': dict(
            add_method='_build_components_sheet',
        ),
        'Buildings': dict(
            add_method='_build_buildings_sheet',
        ),
        'FluidLoops': dict(
            add_method='_build_fluidloops_sheet',
        ),
        'Network': dict(
            add_method='_build_network_sheet',
        ),
        'Options': dict(
            add_method='_build_options_sheet',
        ),
    }

    @classmethod
    def _build_components_sheet(cls, ws):
        """Build the given components Worksheet."""
        # Automatically procure Component classes and their headers,
        # then add them alphabetically.
        headers = cls._find_component_headers()
        for comp_name in sorted(headers):
            ws.append([comp_name])
            ws.append(headers[comp_name])
            ws.append([])
            ws.append([])
        # Formatting.
        n_header_rows = len(headers)
        period = 4
        for i_row in range(1, n_header_rows*period + 1, period):
            for row in ws[f'{i_row}:{i_row+1}']:
                for cell in row:
                    cell.style = 'header'
        ws.column_dimensions['A'].width = 20

    @classmethod
    def _build_buildings_sheet(cls, ws):
        """Build the given buildings Worksheet."""
        # Data.
        ws.append(['IDF file names'])
        # Formatting.
        ws['A1'].style = 'header'
        ws.column_dimensions['A'].width = 15

    @classmethod
    def _build_fluidloops_sheet(cls, ws):
        """Build the given fluid loops Worksheet."""
        # Data.
        headers = ['name', 'node', 'nominal_return_temperature', 'nominal_supply_temperature', 
                    'fluid_capacity', 'fluid_capacitance', 'pump_power_per_kgs', 'temperature_range']
        ws.append(headers)
        # Formatting.
        for cell in ws['1:1']:
            cell.style = 'header'
        widths = [*[9]*2, *[27]*5]
        for cell, width in zip(ws.iter_cols(max_row=1), widths):
            ws.column_dimensions[cell[0].column_letter].width = width

    @classmethod
    def _build_network_sheet(cls, ws):
        """Build the given network Worksheet."""
        headers = ['name', 'network_name', 'equipment', 'buildings', 'latitude', 'longitude', 'time_zone']
        n_connections = 10
        cnxn_headers = ['node_name','line_limit','line_efficiency']
        # Use shorter string "Connexn" to fit number in cell.
        headers.extend([f"Connexn{i}:{h}"
            for i in range(1, n_connections+1)
            for h in cnxn_headers])
        ws.append(headers)
        # Formatting.
        for cell in ws['1:1']:
            cell.style = 'header'
        widths = [*[9]*2, *[28]*3]
        for cell, width in zip(ws.iter_cols(max_row=1), widths):
            ws.column_dimensions[cell[0].column_letter].width = width

    @classmethod
    def _build_timeseriesinputs_sheet(cls, ws):
        """Build the given time series inputs Worksheet."""
        # Data.
        ws.append(['daily', 'nat_gas_price', '<- These are examples.'])
        # Formatting.
        for row in ws['A1:H1']:
            for cell in row:
                cell.style = 'header'
        widths = [13]*2
        for cell, width in zip(ws.iter_cols(max_row=1), widths):
            ws.column_dimensions[cell[0].column_letter].width = width

    @classmethod
    def _build_help_sheet(cls, ws):
        """Build the given help Worksheet."""
        # Formatting.
        ws.sheet_view.showGridLines = False
        # Data.
        ws.add_image(Image('EagersLogo.png'), 'B2')
        # Text.
        text_start_row = 10
        section_header = False
        start_text = False
        with open(EXCEL_INTERFACE_PROJECT_HELP, 'r') as helpfile:
            i_row = text_start_row
            for line in helpfile.readlines():
                line = line.rstrip('\n ')
                if not line:
                    if not start_text:
                        # Text hasn't started yet.
                        continue
                elif line[:2] == '//':
                    # Comment.
                    continue
                elif line[0] == '#':
                    # Section header.
                    line = line.strip('# ')
                    section_header = True
                start_text = True
                cell = ws[f'B{i_row}']
                cell.value = line
                if section_header:
                    for i_col in range(2, 2 + HELP_SECTIONHEADER_WIDTH):
                        ws.cell(i_row, i_col).style = 'help_sectionheader'
                    section_header = False
                i_row += 1


class TestDataTemplateWriter(TestDataTemplateHandler, BaseTemplateWriter):
    """TestData template writer class."""
    pass


class BaseTemplateReader(BaseTemplateHandler):
    """Base template reader class."""

    # Sheets that aren't read directly (or at all).
    # 'TimeSeriesInputs' is read indirectly as a result of reading other
    # sheets.
    # 'Instructions' is just for user reference.
    _ignored_sheets = ('TimeSeriesInputs', 'Instructions')

    # Subclass properties.
    _example_filename: str
    _user_cache_name: str

    def read(self, filename=None):
        """Read user data from filled in template.

        Keyword arguments:
        filename - (str) (Default: None) Custom file name.
        """
        # Load workbook for reading.
        # Use property magic to update the filepath by setting the
        # filename.
        self.filename = filename
        # Use "data_only" flag to avoid reading cell formulas.
        # https://stackoverflow.com/a/35624928/7232335
        if self.filename == self._example_filename:
            wb = load_workbook(
                load_example_file(self._example_filename, self._user_cache_name),
                read_only=True,
                data_only=True,
            )
        else:
            wb = load_workbook(self.filepath, read_only=True, data_only=True)

        # Read sheets.
        workbook_data = {}
        for sheet_name in wb.sheetnames:
            if sheet_name in self._ignored_sheets:
                continue
            workbook_data[sheet_name] = self._read_sheet(wb, sheet_name)
        return workbook_data

    @classmethod
    def read_userfile(cls, filename=None):
        """Convenience method that requires no instantiation. Call
        signature is identical to read().
        """
        reader = cls()
        return reader.read(filename)

    @classmethod
    def _read_sheet(cls, wb, name):
        """Generalized sheet reading operations.

        Positional arguments:
        wb - (Workbook) Workbook to read sheet from.
        name - (str) Name of sheet to read.
        """
        # Generic operations.
        ws = wb[name]
        # Call method for specific operations.
        read_method = cls._sheet_properties[name]['read_method']
        try:
            # Assume the read method is callable.
            return read_method(wb, ws)
        except TypeError:
            # If not callable, find it as an attribute.
            return getattr(cls, read_method)(wb, ws)

    @classmethod
    def _is_empty(cls, row):
        """Check whether the given row is empty."""
        return all(x==None for x in row)

    @classmethod
    def _rstrip_nones(cls, row):
        """Strip None values from the end (right side) of the given row.
        """
        i_startstrip = None
        for i, value in enumerate(reversed(row)):
            if value is None:
                i_startstrip = -i
            else:
                break
        if i_startstrip is None:
            # No trailing Nones values found.
            return row
        return row[:i_startstrip-1]


class ProjectTemplateReader(ProjectTemplateHandler, BaseTemplateReader):
    """Project template reader class."""

    _example_filename = 'default_project.xlsx'
    _sheet_properties = {
        'Components': dict(
            read_method='_read_components_sheet',
        ),
        'Buildings': dict(
            read_method='_read_buildings_sheet',
        ),
        'FluidLoops': dict(
            read_method='_read_fluidloops_sheet',
        ),
        'Network': dict(
            read_method='_read_network_sheet',
        ),
        'Options': dict(
            read_method='_read_optimoptions_sheet',
        ),
        'Reservoirs': dict(
            read_method='_read_reservoir_sheet',
        ),
    }
    _user_cache_name = 'projects'

    def read(self, *args, **kwargs):
        """Overwrite BaseTemplateReader.read()."""
        # First do what the parent method does - read workbook data.
        workbook_data = super().read(*args, **kwargs)
        # Then bundle the project.
        proj_data = dict(
            # name: Remove file suffix, if any.
            name=self.filename.split('.')[0],
            plant=dict(
                generator=workbook_data['Components'],
                building=workbook_data['Buildings'],
                fluid_loop=workbook_data['FluidLoops'],
                network=workbook_data['Network'],
            ),
            options=workbook_data['Options'],
        )
        return proj_data

    @classmethod
    def _read_components_sheet(cls, wb, ws):
        """Read the given components Worksheet."""
        components = []
        type_row = True
        header_row = False
        for row in ws.values:
            if cls._is_empty(row):
                # Blank row.  Next non-empty row contains Component
                # type.
                type_row = True
            elif type_row:
                comp_type = row[0]
                type_row = False
                # Next row contains headers.
                header_row = True
            elif header_row:
                headers = cls._rstrip_nones(row)
                header_row = False
            else:
                comp_info = cls._interpret_components_data(wb, headers, row)
                comp_info.update(_type=comp_type)
                components.append(comp_info)
        return components

    @classmethod
    def _read_buildings_sheet(cls, wb, ws):
        """Read the given buildings Worksheet."""
        values = ws.values
        # First row is header (we know the following rows are file
        # names).
        next(values)
        buildings = []
        for row in values:
            filename = row[0]
            build = bplus_load_building(filename)
            build['filename'] = filename
            buildings.append(build)
        return buildings

    @classmethod
    def _read_fluidloops_sheet(cls, wb, ws):
        """Read the given fluidloops Worksheet."""
        values = ws.values
        headers = cls._rstrip_nones(next(values))
        fluidloops = []
        for row in values:
            row = cls._rstrip_nones(row)
            if row:
                fluidloops.append({
                    key: val
                    for key, val in zip(headers, cls._rstrip_nones(row))
                })
        return fluidloops

    @classmethod
    def _read_network_sheet(cls, wb, ws):
        """Read the given network Worksheet."""
        headers = [cell.value for cell in ws[1]]
        networks = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = cls._rstrip_nones(row)
            if not row:
                # The row has no data in it.
                continue
            net = {
                key: val
                for key, val in zip(headers, row)
                if val is not None
            }
            if 'equipment' in net:
                net['equipment'] = net['equipment'].split(",")
                net['equipment'] = [i.strip() for i in net['equipment']]
            if 'buildings' in net:
                net['buildings'] = net['buildings'].split(",")
                net['buildings'] = [i.strip() for i in net['buildings']]
            networks.append(net)
        return networks

    @classmethod
    def _read_optimoptions_sheet(cls, wb, ws):
        """Read the given optimoptions worksheet."""
        optimoptions = dict(cls._rstrip_nones(x)
            for x in ws.iter_rows(min_row=2, values_only=True))
        # Special handling of start_date value.
        if optimoptions['start_date'] in AUTO_STARTDATE_VALUES:
            optimoptions['start_date'] = None
        else:
            optimoptions['start_date'] = \
                datetime_from_data(optimoptions['start_date'])
        return optimoptions

    @classmethod
    def _read_reservoir_sheet(cls, wb, ws):
        """Read the given reservoir Worksheet."""
        values = ws.values
        headers = cls._rstrip_nones(next(values))
        dams = []
        for row in values:
            row = cls._rstrip_nones(row)
            if row:
                dams.append({
                    key: val
                    for key, val in zip(headers, cls._rstrip_nones(row))
                })
        return dams

    @classmethod
    def _interpret_components_data(cls, wb, headers, data):
        """Return a dictionary of data built from the given headers and
        row of data. This will convert strings to the actual values they
        represent.

        Positional arguments:
        wb - (Workbook) Applicable Workbook.
        headers - (iterable of strings) Header names.
        data - (iterable of data values) Data read from a row of a
            components worksheet.
        """
        # Interpreting function.
        def interpret_value(v):
            # Multiline values: Remove newlines.
            if isinstance(v, str):
                v = v.replace('\n', '')
            # Try using the JSON module to decode values.  This is
            # mostly useful for lists.
            # Can encounter TypeError (if value is not a string) or
            # JSONDecodeError (if value is a string but isn't
            # understood).
            try:
                return json.loads(v)
            except TypeError:
                return v
            except json.JSONDecodeError:
                # Accept "inf" and "-inf" as +/- Infinity.
                if v in ('inf', '-inf'):
                    return float(v)
                return v

        # Build data structure.
        multilevel_field = None
        comp_info = {}
        for h, v in zip(headers, data):
            if '.' in h:
                # NOTE: Assumes only one '.' present, so only 2 levels
                # of compounding are supported.
                field, h = h.rsplit('.', maxsplit=1)
                if multilevel_field != field:
                    comp_info[field] = {}
                info_structure = comp_info[field]
                multilevel_field = field
            else:
                info_structure = comp_info
            if ':' in h and v is not None:
                sheetname, h = h.split(':')
                v_split = v.split(',')
                ts_header, data_header = v_split[0].strip(), v_split[1].strip()
                info = {'timestamp': ts_header, h: data_header}
                # time, values = cls._read_timeseries_input(
                #     wb[sheetname], ts_header, data_header)
                # info = {'timestamp': time, h: values}
            else:
                info = {h: interpret_value(v)}
            if None not in info.values():
                info_structure.update(info)
        return comp_info

    @classmethod
    def _read_timeseries_input(cls, ws, ts_header, data_header):
        """Read a field from a time series worksheet.

        Positional arguments:
        ws - (Worksheet) Applicable Worksheet.
        ts_header - (str) Name of time series header.
        data_header - (str) Name of data header.
        """
        # Get worksheet values generator.
        values = ws.values
        # Read first row of values as headers.
        headers = cls._rstrip_nones(next(values))
        i_ts = headers.index(ts_header)
        i_data = headers.index(data_header)
        # Read remaining rows as data.
        time = []
        data = []
        for row in values:
            time.append(row[i_ts])
            data.append(row[i_data])
        return time, data


class TestDataTemplateReader(TestDataTemplateHandler, BaseTemplateReader):
    """TestData template reader class."""

    _example_filename = 'default_testdata.xlsx'
    _sheet_properties = {
        'NodeData': dict(
            read_method='_read_nodedata_sheet',
        ),
        'Weather': dict(
            read_method='_read_weather_sheet',
        ),
        'Other': dict(
            read_method='_read_other_sheet'
        )
    }
    _user_cache_name = 'test_data'

    def read(self, filename=None):
        """Overwrite BaseTemplateReader.read(). Read user data from
        filled in TestData file template.

        Keyword arguments:
        filename - (str) (Default: None) Custom file name.
        """
        # Load workbook for reading.
        # Use property magic to update the filepath by setting the
        # filename.
        self.filename = filename
        # Use "data_only" flag to avoid reading cell formulas.
        # https://stackoverflow.com/a/35624928/7232335
        if self.filename == self._example_filename:
            wb = load_workbook(
                load_example_file(self._example_filename, self._user_cache_name),
                read_only=True,
                data_only=True,
            )
        else:
            wb = load_workbook(self.filepath, read_only=True, data_only=True)

        # Discover sheets.
        datacat_sheets = [sh for sh in wb.sheetnames
            if sh not in self._ignored_sheets]
        data_categories = [sh.lower() for sh in datacat_sheets]

        # Read sheets for TimeSeriesInputs headers and other info.
        sheet_data = {cat: self._read_sheet(wb, sh)
            for cat, sh in zip(data_categories, datacat_sheets)}

        # Read weather file, if necessary.
        weatherfile_name = sheet_data['weather']['filename']
        weatherfile_designation = sheet_data['weather']['designation']
        weather_data = self._read_weatherfile_data(
            weatherfile_name, weatherfile_designation)
        if weatherfile_name:
            # Delete the weather entry so it isn't used in the upcoming
            # reading of TimeSeriesInputs.
            del sheet_data['weather']

        # Read header information from TimeSeriesInputs sheet.
        table_info, from_tsi = self._read_timeseriesinputs_header_info(
            wb['TimeSeriesInputs'], sheet_data)

        # Add weather info from file, if relevant.
        if weatherfile_name:
            tstamp = weather_data['timestamp']
            ts_resolution = tstamp[1] - tstamp[0]
            designation = weather_data.pop('designation')
            from_tsi['weather'] = False
            table_info['weather'] = {
                TIMESTAMP_RESOLUTION_TABLENAMES[ts_resolution]: dict(
                    designation=designation,
                    data=weather_data,
                    n_row=len(weather_data['timestamp'])
                ),
            }

        # Generate table descriptions.
        table_descriptions = \
            self._generate_table_descriptions(table_info, from_tsi)

        # Generate file names.
        file_info = dict(
            {cat: dict(
                file=(USER_DIR_HDF5_DATASETS / cat
                    / f"{RUNTIME_HDF5_FILE_PREFIX}{cat}{HDF5_SUFFIX}"),
            ) for cat in data_categories}
        )

        # Write data to HDF5.
        self._write_tsi_data_to_hdf5(
            wb['TimeSeriesInputs'],
            table_info,
            file_info,
            table_descriptions,
            from_tsi,
        )

        # Gather all return information.
        for cat in file_info:
            # Remove preamble to file paths.
            file_info[cat]['filename'] = file_info[cat].pop('file').name
            try:
                sheet_data[cat].update(file_info[cat])
            except KeyError:
                sheet_data[cat] = file_info[cat]
        # Flatten.
        return {f'{cat}_{field}': data2
            for cat, data1 in sheet_data.items()
                for field, data2 in data1.items()}

    @classmethod
    def _read_nodedata_sheet(cls, wb, ws):
        """Read the given node data Worksheet."""
        # Get worksheet values generator.
        values = ws.values
        # First row is headers.
        headers = next(values)
        # Read remaining rows as data.
        tsi_headers = {}
        network_info = {}
        for row in values:
            if cls._is_empty(row):
                # Blank row.  Do nothing.
                continue
            else:
                row = cls._rstrip_nones(row)
                if not row[0] is None:
                    net = row[0]
                else:
                    row = list(row)
                    row[0] = net
                if len(row) > 1:
                    # More than just a network name.
                    row_data = dict(zip(headers, row))
                    # Append to set of TimeSeriesInputs headers.
                    for k, v in row_data.items():
                        if ',' not in v:
                            continue
                        tstamp, field = v.replace(' ', '').split(',')
                        try:
                            tsi_headers[tstamp].add(field)
                        except KeyError:
                            tsi_headers[tstamp] = {field}
                        row_data[k] = (tstamp, field)
                    # Append to list of node data.
                    try:
                        network_info[row_data['network']].append(row_data)
                    except KeyError:
                        network_info[row_data['network']] = [row_data]
        return dict(
            tsi_headers=tsi_headers,
            network_info=network_info,
        )

    @classmethod
    def _read_weather_sheet(cls, wb, ws):
        """Read the given weather Worksheet."""
        # Get worksheet values generator.
        values = ws.values
        # Look for file name and designation in first two rows.
        filename = next(values)[1]
        designation = next(values)[1]
        if filename:
            return dict(
                tsi_headers=None, designation=designation, filename=filename)
        # Read remaining rows as data.
        tsi_headers = {}
        tsi_row = 1
        ts_field = None
        for row in values:
            if cls._is_empty(row):
                # Empty row.  Next non-empty row should be timestamp.
                tsi_row = 1
                continue
            elif tsi_row == 1:
                ts_field = row[1]
                tsi_row += 1
            elif ts_field is None:
                # A timestamp field was not specified.
                tsi_row = 0
            elif tsi_row == 2:
                tsi_headers[ts_field] = set(cls._rstrip_nones(row)[1:])
                tsi_row += 1
            elif tsi_row == 3:
                designation = row[1]
                tsi_row += 1
        return dict(
            tsi_headers=tsi_headers,
            designation=designation,
            filename=None,
        )
    @classmethod
    def _read_other_sheet(cls, wb, ws):
        """Read the given node data Worksheet."""
        # Get worksheet values generator.
        values = ws.values
        # First row is headers.
        headers = next(values)
        # Read remaining rows as data.
        tsi_headers = {}
        object_info = {}
        for row in values:
            if cls._is_empty(row):
                # Blank row.  Do nothing.
                continue
            else:
                row = cls._rstrip_nones(row)
                if not row[0] is None:
                    net = row[0]
                else:
                    row = list(row)
                    row[0] = net
                if len(row) > 1:
                    # More than just a object name.
                    row_data = dict(zip(headers, row))
                    # Append to set of TimeSeriesInputs headers.
                    for k, v in row_data.items():
                        if ',' not in v:
                            continue
                        tstamp, field = v.replace(' ', '').split(',')
                        try:
                            tsi_headers[tstamp].add(field)
                        except KeyError:
                            tsi_headers[tstamp] = {field}
                        row_data[k] = (tstamp, field)
                    # Append to list of node data.
                    try:
                        object_info[row_data['object']].append(row_data)
                    except KeyError:
                        object_info[row_data['object']] = [row_data]
        return dict(
            tsi_headers=tsi_headers,
            object_info=object_info,
        )

    @classmethod
    def _read_weatherfile_data(cls, filename, designation):
        if filename:
            try:
                weather_data = read_weather_file(
                    USER_DIR_LOCATION['eplus_weather'] / ensure_suffix(filename, EPW_SUFFIX),
                    designation,
                )
            except FileNotFoundError:
                weather_data = read_weather_file(
                    BUILDPLUS_DIR_PATH /'eplus_weather' / ensure_suffix(filename, EPW_SUFFIX),
                    designation,
                )
            return weather_data

    @classmethod
    def _read_timeseriesinputs_header_info(cls, ws, sheet_data):
        def table_length(ws,col):
            n_pts = 0
            for i,row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if row[col] is None:
                    break
                else:
                    n_pts+=1
            return n_pts
        tsi_headers = cls._rstrip_nones([c.value for c in ws['1']])
        # Indices must be sorted so that read order matches column order
        # in the HDF5 file.
        # TimeSeriesInputs headers are prefixed with "<data category>:"
        # so that it's possible to have the same header name in two
        # different categories.
        tsi_headers_set = set(tsi_headers)
        from_tsi = {}
        table_info = {}
        for cat, data in sheet_data.items():
            from_tsi[cat] = True
            table_info[cat] = {}
            for tablename, hdrs in data.pop('tsi_headers').items():
                h_cat = [tablename, *(f"{cat}:{x}" for x in hdrs)]
                assert set(h_cat) <= tsi_headers_set, \
                    ("Not all named headers appear on the TimeSeriesInputs "
                    "sheet.")
                # Sort headers based on the order in which they appear
                # on the TimeSeriesInputs sheet.
                i_sorted, h_sorted = zip(*sorted(
                    ((i, h) for i, h in enumerate(tsi_headers) if h in h_cat),
                    key=lambda x: x[0],
                ))
                n_pts = table_length(ws,tsi_headers.index(tablename))
                table_info[cat][tablename] = dict(
                    indices=i_sorted,
                    names=h_sorted,
                    n_row = n_pts,
                )
                if cat == 'weather':
                    table_info[cat][tablename]['designation'] = \
                        sheet_data['weather'].pop('designation')
        return table_info, from_tsi

    @classmethod
    def _generate_table_descriptions(cls, table_info, from_tsi):
        table_descriptions = {}
        for cat, data in table_info.items():
            table_descriptions[cat] = {}
            if from_tsi[cat]:
                for table, h_info in data.items():
                    table_descriptions[cat][table] = {}
                    dtype = []
                    for h in h_info['names']:
                        h_split = h.split(':')
                        try:
                            h = h_split[1]
                        except IndexError:
                            h = 'timestamp'
                        dtype.append((h, 'f8'))
                    table_descriptions[cat][table] = np.dtype(dtype)
            else:
                for table, h_info in data.items():
                    table_descriptions[cat][table] = {}
                    dtype = []
                    for h in h_info['data']:
                        dtype.append((h, 'f8'))
                    table_descriptions[cat][table] = np.dtype(dtype)
        return table_descriptions

    @classmethod
    def _gather_data_rows(cls, ws, n_pts, i_headers, dtype):
        """Returns a structured array of data values for a given set of
        headers.
        """
        data = np.empty(dtype=dtype, shape=n_pts)
        timestamp = np.empty(dtype='O', shape=n_pts)
        i_headers = np.array(i_headers)
        non_ts = np.array(dtype.names) != 'timestamp'
        i_non_ts = i_headers[non_ts]
        i_ts = i_headers[np.invert(non_ts)][0]
        # ReadOnlyWorksheet does not have iter_cols().  Use
        # iter_rows().
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row = n_pts+1, values_only=True)):
            row = cls._rstrip_nones(row)
            data[i] = (0, *(row[j] for j in i_non_ts))
            timestamp[i] = datetime_from_data(row[i_ts])
        # Convert datetimes to floats; this is the only way PyTables
        # can store datetime values.
        data['timestamp'] = DFC.d2f_arr2arr(timestamp)
        return data

    @classmethod
    def _write_tsi_data_to_hdf5(cls,
            ws, table_info, file_info, table_descriptions, from_tsi):
        for cat, data in table_info.items():
            new_hdf5_file(file_info[cat]['file'], overwrite=True)
            for table, h_info in data.items():
                if from_tsi[cat]:
                    records = cls._gather_data_rows(ws,h_info['n_row'],h_info['indices'],table_descriptions[cat][table])
                else:
                    records = np.empty(
                        dtype=table_descriptions[cat][table],
                        shape=h_info['n_row'],
                    )
                    for k, v in h_info['data'].items():
                        if k == 'timestamp':
                            v = DFC.d2f_lst2arr(v)
                        records[k] = v
                write_records_to_new_hdf5_table(
                    file_info[cat]['file'],
                    table,
                    records,
                )
                if cat == 'weather':
                    # Write designation to file.
                    write_user_attributes(
                        file_info[cat]['file'],
                        table,
                        dict(designation=h_info['designation']),
                    )
